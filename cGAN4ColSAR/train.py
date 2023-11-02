from __future__ import print_function
import argparse
import os
from math import log10

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
import torch.backends.cudnn as cudnn

from networks import define_G, define_D, GANLoss, get_scheduler, update_learning_rate
from dataset import *

import time
from datetime import datetime
import sarcolor_metrics as mtc
from utils import *
import openpyxl
import numpy as np
from tqdm import tqdm

# Training settings
parser = argparse.ArgumentParser(description='SAR Colorization')
parser.add_argument('--method', type=str, default='CAN4ColSAR')
parser.add_argument('--dataset', type=str, default='sen12ms_cr')
parser.add_argument('--batch_size', type=int, default=8, help='training batch size')
parser.add_argument('--test_batch_size', type=int, default=1, help='testing batch size')
parser.add_argument('--input_nc', type=int, default=1, help='input image channels')
parser.add_argument('--output_nc', type=int, default=3, help='output image channels')
parser.add_argument('--ngf', type=int, default=64, help='generator filters in first conv layer')
parser.add_argument('--ndf', type=int, default=64, help='discriminator filters in first conv layer')
parser.add_argument('--netD', type=str, default='basic',help='specify discriminator architecture [basic | n_layers | pixel]. The basic model is a 70x70 PatchGAN. n_layers allows you to specify the layers in the discriminator')
parser.add_argument('--netG', type=str, default='unet_128',help='specify generator architecture [unet_256 | unet_128 | unet_64]')
parser.add_argument('--epoch_count', type=int, default=1, help='the starting epoch count')
parser.add_argument('--niter', type=int, default=300, help='# of iter at starting learning rate')
parser.add_argument('--niter_decay', type=int, default=0, help='# of iter to linearly decay learning rate to zero')
parser.add_argument('--lr', type=float, default=0.0001, help='initial learning rate for adam')
parser.add_argument('--lr_policy', type=str, default='lambda', help='learning rate policy: lambda|step|plateau|cosine')
parser.add_argument('--lr_decay_iters', type=int, default=50, help='multiply by a gamma every lr_decay_iters iterations')
parser.add_argument('--beta1', type=float, default=0.5, help='beta1 for adam. default=0.5')
parser.add_argument('--threads', type=int, default=0, help='number of threads for data loader to use')  # default=4
parser.add_argument('--seed', type=int, default=123, help='random seed to use. Default=123')
parser.add_argument('--lamb', type=int, default=210, help='weight on L1 term in objective')
parser.add_argument('--log_freq', type=int, default=200, help='print the loss every abc iterations')
parser.add_argument('--ckpt_freq', type=int, default=20, help='save the model every abc epochs')
parser.add_argument('--gpu_id', type=str, default='0')
parser.add_argument('--eval_epoch', type=int, default= 300)
opt = parser.parse_args()
print(opt)
cudnn.benchmark = True

version = 1
root_path = "../SEN12MS_CR_SARColorData/"
imgsave_dir = './result/test_samples_v%s/' % version
checkpoint_dir = "./result/checkpoint_v%s" % version
record_dir = './result/record_v%s/' % version
eval_ckpt_path = "./result/checkpoint_v%s/netG_model_epoch_%s.pth" % (version, opt.eval_epoch)

if not os.path.exists(checkpoint_dir):
    os.makedirs(checkpoint_dir)
if not os.path.exists(imgsave_dir):
    os.makedirs(imgsave_dir)
if not os.path.exists(record_dir):
    os.makedirs(record_dir)

print('===> Loading datasets')
train_set = DatasetFromFolder(root_path, "train")
test_set = DatasetFromFolder(root_path, "test")

training_data_loader = DataLoader(dataset=train_set, num_workers=opt.threads, batch_size=opt.batch_size, shuffle=True)
testing_data_loader = DataLoader(dataset=test_set, num_workers=opt.threads, batch_size=opt.test_batch_size, shuffle=False)

## Device configuration
os.environ['CUDA_DEVICE_ORDER'] = 'PCI_BUS_ID'
os.environ['CUDA_VISIBLE_DEVICES'] = opt.gpu_id
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
print(device)

print('===> Building models')
net_g = define_G(opt.input_nc, opt.output_nc, opt.ngf, opt.netG,'batch', False, 'normal', 0.02)
net_d = define_D(opt.input_nc + opt.output_nc, opt.ndf, opt.netD, n_layers_D=3, norm='batch', init_type='normal', init_gain=0.02)
net_g, net_d = net_g.to(device),net_d.to(device)

if (torch.cuda.device_count() > 1):
    print("Let's use", torch.cuda.device_count(), "GPUs!")
    net_g = nn.DataParallel(net_g)
    net_d = nn.DataParallel(net_d)

criterionGAN = GANLoss().to(device)
criterionL1 = nn.L1Loss().to(device)
criterionMSE = nn.MSELoss().to(device)

# setup optimizer
optimizer_g = optim.Adam(net_g.parameters(), lr=opt.lr, betas=(opt.beta1, 0.999))
optimizer_d = optim.Adam(net_d.parameters(), lr=opt.lr, betas=(opt.beta1, 0.999))
net_g_scheduler = get_scheduler(optimizer_g, opt)
net_d_scheduler = get_scheduler(optimizer_d, opt)

print('Training Begin: ', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
start = time.time()

train_loss_record = open('%s/train_loss_record.txt' % record_dir, "w")
test_sam_record = open('%s/test_sam_record.txt' % record_dir, "w")

early_stopping = EarlyStopping(patience=opt.patience, verbose=True, path=eval_ckpt_path)
for epoch in range(opt.epoch_count, opt.niter + opt.niter_decay + 1):
    # train
    for iteration, batch in tqdm(enumerate(training_data_loader, 1),total=len(training_data_loader)):
        # forward
        real_a, real_b = batch[0].to(device), batch[1].to(device)
        fake_b = net_g(real_a)

        ######################
        # (1) Update D network
        ######################

        optimizer_d.zero_grad()
        
        # train with fake
        fake_ab = torch.cat((real_a, fake_b), 1)
        pred_fake = net_d.forward(fake_ab.detach())
        loss_d_fake = criterionGAN(pred_fake, False)  # False，target_tensor = self.fake_label

        # train with real
        real_ab = torch.cat((real_a, real_b), 1)
        pred_real = net_d.forward(real_ab)
        loss_d_real = criterionGAN(pred_real, True)  # True，target_tensor = self.real_label
        
        # Combined D loss
        loss_d = (loss_d_fake + loss_d_real) * 0.5

        loss_d.backward()
       
        optimizer_d.step()

        ######################
        # (2) Update G network
        ######################

        optimizer_g.zero_grad()

        # First, G(A) should fake the discriminator
        fake_ab = torch.cat((real_a, fake_b), 1)
        pred_fake = net_d.forward(fake_ab)
        loss_g_gan = criterionGAN(pred_fake, True)

        # Second, G(A) = B
        loss_g_l1 = criterionL1(fake_b, real_b) * opt.lamb
        loss_g = loss_g_gan + loss_g_l1
        
        loss_g.backward()

        optimizer_g.step()

        if iteration % opt.log_freq == 0:
            print("===> Epoch[{}/{}]({}/{}): Loss_D: {:.4f} Loss_G: {:.4f} Loss_g_gan: {:.4f} Loss_g_L1: {:.4f}".format(
                epoch, opt.niter + opt.niter_decay, iteration, len(training_data_loader), loss_d.item(), loss_g.item(), loss_g_gan.item(), loss_g_l1.item()))
            train_loss_record.write("===> Epoch[{}/{}]({}/{}): Loss_D: {:.4f} Loss_G: {:.4f} Loss_gan: {:.4f} Loss_g_L1: {:.4f}\n".format(
                epoch, opt.niter + opt.niter_decay, iteration, len(training_data_loader), loss_d.item(), loss_g.item(), loss_g_gan.item(), loss_g_l1.item()))

    update_learning_rate(net_g_scheduler, optimizer_g)
    update_learning_rate(net_d_scheduler, optimizer_d)

    #checkpoint
    if epoch % opt.ckpt_freq == 0:
        print('Save chckpoint model epoch%s' % epoch)
        torch.save(net_g.state_dict(), os.path.join(checkpoint_dir,'netG_model_epoch_%s.pth' % epoch))

train_loss_record.close()


def eval(model, dataloader, imgsave_dir):
    model.load_state_dict(torch.load(eval_ckpt_path))
    model.eval()

    NRMSE, SAM = [], []
    print('Eval model ... Save image...')
    for iteration, batch in enumerate(dataloader):

        sar, gt, sarcolor_filename = batch[0].to(device), batch[1].to(device), batch[2]
        out_img = model(sar)

        out_img = out_img * 2048 + 2048
        gt = gt * 2048 + 2048
        out_img = trim_image(out_img, L=0, R=2 ** 12 - 1)
        gt = trim_image(gt, L=0, R=2 ** 12 - 1)

        gt, out_img = gt.detach().cpu().numpy(), out_img.detach().cpu().numpy()
        gt, out_img = np.transpose(gt, (0, 2, 3, 1)), np.transpose(out_img, (0, 2, 3, 1))
        gt, out_img = gt[0], out_img[0]

        save_image(os.path.join(imgsave_dir, sarcolor_filename[0]), out_img.transpose(2,0,1), 3)

        NRMSE.append(mtc.NRMSE_numpy(gt,out_img, norm_type = 'euclidean'))
        SAM.append(mtc.SAM_numpy(gt,out_img, sewar=True))

    print("NRMSE，SAM")
    print("%.4lf±%.4lf, %.4lf±%.4lf" %
          (np.mean(NRMSE), np.std(NRMSE), np.mean(SAM), np.std(SAM)))

    wb_eval_metric = openpyxl.Workbook()
    ws_eval_metric = wb_eval_metric.create_sheet('sheet1',0)

    metrics_name_list = ["NRMSE", "SAM"]
    metrics_list = ["%.4lf±%.4lf" % (np.mean(NRMSE), np.std(NRMSE)),
                    "%.4lf±%.4lf" % (np.mean(SAM), np.std(SAM))]

    for i in range(len(metrics_name_list)):
        ws_eval_metric.cell(row=1, column=i + 1).value = metrics_name_list[i]
        ws_eval_metric.cell(row=2, column=i + 1).value = metrics_list[i]

    wb_eval_metric.save('%s/%s_eval_metric_record.xlsx' % (record_dir,opt.method))

eval(net_g, testing_data_loader, imgsave_dir)

print('==>Total time consuming： {:.2f}h\n'.format((time.time() - start) / 3600))
print('End Time: ', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
